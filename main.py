import openpyxl
import cv2


def ndarray_to_rgb(ndarray):
    re = ''
    for c in ndarray:
        re += f'{int(c):02x}'
    return re.upper()


imgfn = 'input.jpg'
img = cv2.imread(imgfn, cv2.IMREAD_COLOR)[:, :, ::-1]


y, x, _ = img.shape
fxy = 800/y
if fxy < 1:
    img = cv2.resize(img, (0, 0), fx=fxy, fy=fxy,
                     interpolation=cv2.INTER_NEAREST)


wb = openpyxl.Workbook()
ws01 = wb['Sheet']
pix = 10
lng = len(img)
for row in range(lng):
    print('.', end='')
    for col in range(len(img[row])):
        fgColor = ndarray_to_rgb(img[row][col])
        pattern_fill = openpyxl.styles.PatternFill('solid', fgColor=fgColor)
        ws01.cell(row=row+1, column=col+1, value=None).fill = pattern_fill

        ws01.column_dimensions[openpyxl.utils.get_column_letter(
            col+1)].width = pix*254.86/1789

    ws01.row_dimensions[row+1].height = pix*409.50/546


excelfn = 'output.xlsx'
print(f'save to {excelfn}')
wb.save(excelfn)
